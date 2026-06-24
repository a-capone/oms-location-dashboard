from __future__ import annotations

import argparse
import json
import os
import re
import urllib.parse
import urllib.request
import warnings
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import paramiko
from google.cloud import bigquery
from openpyxl import load_workbook


warnings.filterwarnings("ignore", message="Workbook contains no default style.*", category=UserWarning)


PROJECT = "tlg-business-intelligence-prd"
VIEW_ID = "tlg-business-intelligence-prd.til.v_oms_active_location_dashboard"

DEFAULT_TEMPLATE = Path("templates/dashboard.html")
DEFAULT_REMOTE_DIRS = ["/report/Outbound/Uscite", "/report/Outbound"]
DEFAULT_TOKEN_URL = "https://integrations.thelevelgroup.com/identity/oauth2/token"
DEFAULT_INVENTORY_URL = "https://integrations.thelevelgroup.com/invapp-dab-prod/v2/InventoryWmsMonitoring"
INVENTORY_HISTORY_LIMIT = 168

BRAND_CONFIGS: dict[str, dict[str, Any]] = {
    "FE": {"name": "Ferrari",         "brand_code": 48, "temp": "TLTEMP0048", "gxo": "TLITGX0048", "wh": "TLITWH0048"},
    "DG": {"name": "Dolce&Gabbana",   "brand_code": 15, "temp": "TLTEMP0015", "gxo": "TLITGX0015", "wh": "TLITWH0015"},
    "DJ": {"name": "La DoubleJ",      "brand_code": 11, "temp": "TLTEMP0011", "gxo": "TLITGX0011", "wh": "TLITWH0011"},
    "BB": {"name": "Brooks Brothers", "brand_code": 42, "temp": "TLTEMP0042", "gxo": "TLITGX0042", "wh": "BBITST0001"},
    "WL": {"name": "FrankBros",       "brand_code": 12, "temp": "TLTEMP0012", "gxo": "TLITGX0012", "wh": "TLITWH0012"},
}

GEODIS_STATUS = {
    "1": "INTEGRATO",
    "2": "IN ATTESA DI PICKING/EVASO CON ROTTURE DI STOCK",
    "3": "PICKING",
    "4": "PICKING",
    "5": "PACKING",
    "7": "PACKED (MA NON BORDERIZZATO)",
    "8": "SHIPPED",
}

GEODIS_SORT = {
    "1": 10,
    "2": 20,
    "3": 30,
    "4": 40,
    "5": 50,
    "7": 70,
    "8": 80,
}


@dataclass
class GeodisFile:
    remote_dir: str
    filename: str
    size: int
    mtime: int

    @property
    def remote_path(self) -> str:
        return f"{self.remote_dir.rstrip('/')}/{self.filename}"


def normalize_key(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    return text.upper()


def normalize_status(value: Any) -> str | None:
    key = normalize_key(value)
    if not key:
        return None
    match = re.search(r"\d+", key)
    return match.group(0) if match else None


def filename_sort_key(name: str) -> tuple[int, ...]:
    match = re.search(r"Uscite_(\d{4})_(\d{2})_(\d{2})(?:_(\d{2})_(\d{2})_(\d{2}))?", name, re.IGNORECASE)
    if not match:
        return (0,)
    return tuple(int(part or 0) for part in match.groups())


def fetch_bigquery_rows(locations: list[str]) -> list[dict[str, Any]]:
    client = bigquery.Client(project=PROJECT)
    query = f"""
    SELECT
      dashboard_refreshed_at_utc,
      assigned_location_code,
      location_label,
      order_number,
      external_order_id,
      shipment_order_number,
      oms_order_id,
      oms_shipment_number,
      brand,
      shipment_status,
      fulfillment_status,
      workflow_shipment_state,
      active_task_name,
      active_task_id,
      order_submit_datetime_rome,
      shipment_creation_datetime_rome,
      update_datetime_rome,
      hours_since_update,
      days_since_update,
      days_since_creation,
      update_age_bucket,
      dashboard_state,
      attention_bucket,
      attention_sort,
      is_canceled,
      is_ready
    FROM `{VIEW_ID}`
    WHERE shipment_status != 'CANCELED'
      AND assigned_location_code IN UNNEST(@locations)
    ORDER BY attention_sort, days_since_update DESC, assigned_location_code, external_order_id, oms_shipment_number
    """
    job_config = bigquery.QueryJobConfig(
        query_parameters=[bigquery.ArrayQueryParameter("locations", "STRING", locations)]
    )
    rows: list[dict[str, Any]] = []
    for row in client.query(query, job_config=job_config).result():
        out: dict[str, Any] = {}
        for key, value in dict(row.items()).items():
            if hasattr(value, "isoformat"):
                value = value.isoformat()
            out[key] = value
        rows.append(out)
    return rows


def connect_sftp() -> paramiko.SFTPClient:
    host = os.environ.get("GEODIS_SFTP_HOST", "stnetlglog.blob.core.windows.net")
    user = os.environ.get("GEODIS_SFTP_USER", "stnetlglog.geodis")
    port = int(os.environ.get("GEODIS_SFTP_PORT", "22"))
    password = os.environ.get("GEODIS_SFTP_PASSWORD")
    if not password:
        raise RuntimeError("Missing GEODIS_SFTP_PASSWORD environment variable")

    transport = paramiko.Transport((host, port))
    transport.connect(username=user, password=password)
    sftp = paramiko.SFTPClient.from_transport(transport)
    sftp._transport_ref = transport  # type: ignore[attr-defined]
    return sftp


def close_sftp(sftp: paramiko.SFTPClient) -> None:
    transport = getattr(sftp, "_transport_ref", None)
    sftp.close()
    if transport:
        transport.close()


def find_latest_geodis_file(sftp: paramiko.SFTPClient) -> GeodisFile:
    configured = os.environ.get("GEODIS_SFTP_DIR")
    remote_dirs = [configured] if configured else DEFAULT_REMOTE_DIRS
    found: list[GeodisFile] = []
    errors: list[str] = []

    for remote_dir in [d for d in remote_dirs if d]:
        try:
            entries = sftp.listdir_attr(remote_dir)
        except Exception as exc:
            errors.append(f"{remote_dir}: {str(exc).split(' - RequestId:')[0]}")
            continue
        for entry in entries:
            name = entry.filename
            if not name.lower().startswith("uscite_"):
                continue
            if not name.lower().endswith((".xlsx", ".xlsm", ".xls")):
                continue
            found.append(GeodisFile(remote_dir, name, int(entry.st_size), int(entry.st_mtime)))

    if not found:
        detail = "; ".join(errors) if errors else "no matching Uscite_*.xlsx files"
        raise RuntimeError(f"No Geodis Uscite Excel found ({detail})")

    return sorted(found, key=lambda item: (filename_sort_key(item.filename), item.mtime, item.filename), reverse=True)[0]


def download_geodis_file(sftp: paramiko.SFTPClient, remote_file: GeodisFile, cache_dir: Path) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    local_path = cache_dir / remote_file.filename
    if local_path.exists() and local_path.stat().st_size == remote_file.size:
        return local_path
    sftp.get(remote_file.remote_path, str(local_path))
    return local_path


def parse_geodis_excel(path: Path, source_file: GeodisFile) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = ["ord"] if "ord" in workbook.sheetnames else workbook.sheetnames
    index: dict[str, dict[str, Any]] = {}

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        if hasattr(sheet, "reset_dimensions"):
            sheet.reset_dimensions()
        headers: list[str] | None = None

        for row_values in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value).strip() for value in row_values]
            if headers is None:
                if any(value.upper() == "STATO_OC" for value in values):
                    headers = values
                continue

            if not any(values):
                continue

            record = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
            status_code = normalize_status(record.get("STATO_OC"))
            if not status_code:
                continue

            status_label = GEODIS_STATUS.get(status_code, f"STATO GEODIS {status_code}")
            item = {
                "geodis_found": True,
                "geodis_status_code": status_code,
                "geodis_status_label": status_label,
                "geodis_sort": GEODIS_SORT.get(status_code, 0),
                "geodis_reference": None,
                "geodis_reference_field": None,
                "geodis_riferimento_oc": record.get("RIFERIMENTO_OC") or None,
                "geodis_riferimento_cliente": record.get("RIFERIMENTO_CLIENTE") or None,
                "geodis_ordine_geode": record.get("ORDINE_GEODE") or None,
                "geodis_brand_code": record.get("BRAND") or None,
                "geodis_source_file": source_file.filename,
                "geodis_source_dir": source_file.remote_dir,
                "geodis_sheet": sheet_name,
            }

            for field in ("RIFERIMENTO_CLIENTE", "RIFERIMENTO_OC", "ORDINE_GEODE"):
                key = normalize_key(record.get(field))
                if not key:
                    continue
                candidate = dict(item)
                candidate["geodis_reference"] = key
                candidate["geodis_reference_field"] = field
                existing = index.get(key)
                if existing is None or candidate["geodis_sort"] >= existing.get("geodis_sort", 0):
                    index[key] = candidate

    return index


def enrich_with_geodis(rows: list[dict[str, Any]], geodis_index: dict[str, dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    matched = 0
    for row in rows:
        match: dict[str, Any] | None = None
        match_field: str | None = None
        for field in ("shipment_order_number", "external_order_id", "order_number", "oms_shipment_number"):
            key = normalize_key(row.get(field))
            if key and key in geodis_index:
                match = geodis_index[key]
                match_field = field
                break

        if match:
            row.update(match)
            row["geodis_match_field"] = match_field
            matched += 1
        else:
            row.update({
                "geodis_found": False,
                "geodis_status_code": None,
                "geodis_status_label": "Non presente in file Geodis",
                "geodis_sort": 0,
                "geodis_reference": None,
                "geodis_reference_field": None,
                "geodis_match_field": None,
                "geodis_riferimento_oc": None,
                "geodis_riferimento_cliente": None,
                "geodis_ordine_geode": None,
                "geodis_brand_code": None,
                "geodis_source_file": None,
                "geodis_source_dir": None,
                "geodis_sheet": None,
            })
    return rows, matched


def fetch_inventory_rows() -> dict[str, Any]:
    client_id = os.environ.get("INVAPP_CLIENT_ID")
    client_secret = os.environ.get("INVAPP_CLIENT_SECRET")
    if not client_id or not client_secret:
        raise RuntimeError("Missing INVAPP_CLIENT_ID or INVAPP_CLIENT_SECRET environment variable")

    locations = os.environ.get("INVAPP_LOCATIONS", DEFAULT_INVENTORY_LOCATIONS)
    token_url = os.environ.get("INVAPP_TOKEN_URL", DEFAULT_TOKEN_URL)
    inventory_url = os.environ.get("INVAPP_INVENTORY_URL", DEFAULT_INVENTORY_URL)

    token_body = urllib.parse.urlencode({
        "client_id": client_id,
        "client_secret": client_secret,
    }).encode("utf-8")
    token_request = urllib.request.Request(
        token_url,
        data=token_body,
        method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
    )
    with urllib.request.urlopen(token_request, timeout=60) as response:
        token_payload = json.loads(response.read().decode("utf-8"))
    token = token_payload.get("access_token")
    if not token:
        raise RuntimeError("Inventory auth response did not include access_token")

    inventory_body = json.dumps({"locations": locations}).encode("utf-8")
    inventory_request = urllib.request.Request(
        inventory_url,
        data=inventory_body,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {token}",
        },
    )
    with urllib.request.urlopen(inventory_request, timeout=120) as response:
        inventory_payload = json.loads(response.read().decode("utf-8"))

    raw_rows = inventory_payload.get("value", []) if isinstance(inventory_payload, dict) else []
    if not isinstance(raw_rows, list):
        raw_rows = []

    rows: list[dict[str, Any]] = []
    for item in raw_rows:
        if not isinstance(item, dict):
            continue
        quantity = item.get("quantity")
        try:
            quantity_value = int(quantity)
        except (TypeError, ValueError):
            quantity_value = None
        rows.append({
            "locationName": item.get("locationName"),
            "quantity": quantity_value,
            "updatedAt": item.get("updatedAt"),
        })

    return {
        "locations": locations,
        "fetchedAtUtc": datetime.now(timezone.utc).isoformat(),
        "rowCount": len(rows),
        "rows": rows,
    }


def inventory_quantity_map(inventory: dict[str, Any]) -> dict[str, int | None]:
    out: dict[str, int | None] = {}
    for row in inventory.get("rows", []):
        if not isinstance(row, dict):
            continue
        location = row.get("locationName")
        if not location:
            continue
        quantity = row.get("quantity")
        try:
            out[str(location)] = int(quantity)
        except (TypeError, ValueError):
            out[str(location)] = None
    return out


def extract_payload_from_html(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    html = path.read_text(encoding="utf-8", errors="ignore")
    match = re.search(r"window\.OMS_DASHBOARD_DATA = (.*?);\s*</script>", html, flags=re.DOTALL)
    if not match:
        return None
    try:
        return json.loads(match.group(1))
    except json.JSONDecodeError:
        return None


def previous_inventory_history(output: Path) -> list[dict[str, Any]]:
    candidates = []
    previous_html = os.environ.get("DASHBOARD_PREVIOUS_HTML")
    if previous_html:
        candidates.append(Path(previous_html))
    candidates.append(output)

    for candidate in candidates:
        payload = extract_payload_from_html(candidate)
        if not payload:
            continue
        history = ((payload.get("inventory") or {}).get("history") or [])
        if isinstance(history, list):
            return [item for item in history if isinstance(item, dict)]
    return []


def append_inventory_history(history: list[dict[str, Any]], inventory: dict[str, Any], brand_cfg: dict[str, Any]) -> list[dict[str, Any]]:
    quantities = inventory_quantity_map(inventory)
    temp_loc = brand_cfg["temp"]
    gxo_loc = brand_cfg["gxo"]
    wh_loc = brand_cfg["wh"]
    point: dict[str, Any] = {
        "fetchedAtUtc": inventory.get("fetchedAtUtc"),
        temp_loc: quantities.get(temp_loc),
        gxo_loc: quantities.get(gxo_loc),
    }
    if wh_loc:
        point[wh_loc] = quantities.get(wh_loc)
    if not point["fetchedAtUtc"]:
        return history[-INVENTORY_HISTORY_LIMIT:]

    cleaned = [item for item in history if item.get("fetchedAtUtc") != point["fetchedAtUtc"]]
    cleaned.append(point)
    cleaned.sort(key=lambda item: str(item.get("fetchedAtUtc") or ""))
    return cleaned[-INVENTORY_HISTORY_LIMIT:]


def render_template(template: Path, output: Path, payload: dict[str, Any]) -> None:
    html = template.read_text(encoding="utf-8")
    data = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    new_html = html.replace("__OMS_DATA__", data, 1)
    if new_html == html:
        raise RuntimeError(f"Placeholder __OMS_DATA__ not found in template {template}")
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(new_html, encoding="utf-8")


def build_payload(
    rows: list[dict[str, Any]],
    geodis_file: GeodisFile,
    geodis_index_size: int,
    geodis_matches: int,
    inventory: dict[str, Any],
    inventory_history: list[dict[str, Any]],
    brand_cfg: dict[str, Any],
) -> dict[str, Any]:
    inventory = dict(inventory)
    inventory["history"] = inventory_history
    return {
        "generatedAtUtc": datetime.now(timezone.utc).isoformat(),
        "source": VIEW_ID,
        "rowCount": len(rows),
        "lastHourlyRefreshAtUtc": datetime.now(timezone.utc).isoformat(),
        "brandConfig": brand_cfg,
        "geodis": {
            "fileName": geodis_file.filename,
            "remoteDir": geodis_file.remote_dir,
            "remotePath": geodis_file.remote_path,
            "fileSize": geodis_file.size,
            "fileMtime": geodis_file.mtime,
            "indexedReferences": geodis_index_size,
            "matchedShipments": geodis_matches,
            "statusMap": GEODIS_STATUS,
        },
        "inventory": inventory,
        "rows": rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the mono-file OMS dashboard with BigQuery and Geodis SFTP Excel data.")
    parser.add_argument("--brand", required=True, choices=list(BRAND_CONFIGS.keys()), help="Brand to build the dashboard for (e.g. FE, DG, DJ, BB, WL)")
    parser.add_argument("--output", default=None, help="Path to HTML output, default tmp/{brand}/index.html")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="Path to HTML template, default templates/dashboard.html")
    parser.add_argument("--cache-dir", default="runtime/geodis", help="Local cache for downloaded Geodis Excel files")
    args = parser.parse_args()

    brand_cfg = BRAND_CONFIGS[args.brand]
    locations = [brand_cfg["temp"], brand_cfg["gxo"]]
    if brand_cfg["wh"]:
        locations.append(brand_cfg["wh"])

    output = Path(args.output) if args.output else Path(f"tmp/{args.brand.lower()}/index.html")
    template = Path(args.template)
    if not template.exists():
        raise RuntimeError(f"Template not found: {template}")

    rows = fetch_bigquery_rows(locations)
    prior_inventory_history = previous_inventory_history(output)

    sftp = connect_sftp()
    try:
        latest = find_latest_geodis_file(sftp)
        local_file = download_geodis_file(sftp, latest, Path(args.cache_dir))
    finally:
        close_sftp(sftp)

    geodis_index = parse_geodis_excel(local_file, latest)
    enriched, matches = enrich_with_geodis(rows, geodis_index)
    inventory = fetch_inventory_rows()
    inventory_history = append_inventory_history(prior_inventory_history, inventory, brand_cfg)
    payload = build_payload(enriched, latest, len(geodis_index), matches, inventory, inventory_history, brand_cfg)
    render_template(template, output, payload)

    print(json.dumps({
        "brand": args.brand,
        "output": str(output),
        "rows": len(enriched),
        "geodis_file": latest.filename,
        "geodis_references": len(geodis_index),
        "geodis_matched_shipments": matches,
        "inventory_locations": inventory["locations"],
        "inventory_rows": inventory["rowCount"],
        "inventory_history_points": len(inventory_history),
    }, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
